"""Parse archived 'Menu' daily workbooks into a structured picks database.

Input: scratchpad/menu_raw/<sheet_id>.xlsx (bulk-fetched daily snapshots)
       scratchpad/menu_archive_index_full.json (sheet_id -> date)
Output: data/menu_picks.csv — one row per pick cell:
        date, league, category, text, line, gt, apt, sheet_id

Layout: category blocks announced by a header row (SHARP MONEY / BOOK NEEDS /
SQUARES / RLM / WHALE / MODEL BEST VALUES / MEGA RLM / MEGA SHARPS / SHARP
PROPS), league sub-headers (MLB/NBA/...) per column, pick rows below. Format
drifts across eras — detection is contains-based and per-column.

Research artifact for the market-signal predictivity study. Market data —
NEVER a model input (constitution); evaluation/diagnostics only.
"""
from __future__ import annotations

import csv
import json
import os
import re
import sys

import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

SP = (r"C:/Users/Admin/AppData/Local/Temp/claude/"
      r"C--Users-Admin-Projects-Sports-Odds-Algorithms/"
      r"c82ef32b-d2a5-4470-9d81-a522f1e0284f/scratchpad")
RAW = SP + "/menu_raw"

CATS = [
    ("mega_sharps", ("mega sharp",)),
    ("mega_rlm", ("mega rlm",)),
    ("sharp_props", ("sharp prop",)),
    ("sharp", ("sharp money",)),
    ("book_needs", ("book need", "sportsbook need", "what the sportsbook")),
    ("squares", ("square",)),
    ("rlm", ("reverse line",)),
    ("whale", ("whale",)),
    ("model_best", ("model best",)),
    ("tier1", ("tier 1",)),
    ("big_bets", ("biggest bets",)),
    ("dog_plays", ("best dog",)),
]
LEAGUES = {"MLB": "MLB", "NBA": "NBA", "WNBA": "WNBA", "NHL": "NHL",
           "CBB": "CBB", "NCAAB": "CBB", "CFB": "CFB", "NCAAF": "CFB",
           "NFL": "NFL", "CFL": "CFL", "UFC": "UFC", "SOCCER": "SOC",
           "NBA / WNBA": "NBA", "EPL": "SOC", "ATP": "TEN", "WTA": "TEN"}

def cat_of(txt):
    t = txt.lower()
    for name, keys in CATS:
        if any(k in t for k in keys):
            return name
    return None

def league_of(txt):
    t = re.sub(r"[^A-Z/ ]", "", txt.upper()).strip()
    return LEAGUES.get(t)

def parse_book(path, date, sid, out):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    tab = next((s for s in wb.sheetnames if s.strip().lower() == "todays menu"), None)
    if tab is None:
        tab = wb.sheetnames[0]
    ws = wb[tab]
    col_cat = {}         # col -> category  (from most recent category header row)
    col_lg = {}          # col -> league    (from most recent league row)
    gt_col, apt_col = None, None
    n0 = len(out)
    for row in ws.iter_rows(max_row=200, max_col=24):
        vals = [(c.column, str(c.value).strip()) for c in row
                if c.value is not None and str(c.value).strip()]
        if not vals:
            continue
        texts = [v for _, v in vals]
        cats_here = [(col, cat_of(v)) for col, v in vals if cat_of(v)]
        lgs_here = [(col, league_of(v)) for col, v in vals if league_of(v)]
        # header row: at least one category name -> reset category map region
        if cats_here and len(cats_here) >= max(1, len([1 for _, v in vals if len(v) > 3]) // 3):
            col_cat = {}
            for col, cat in cats_here:
                col_cat[col] = cat
            continue
        # league row: GT/APT markers or multiple league names
        if lgs_here and (len(lgs_here) > 1
                         or any(v.upper() in ("GT", "GAME TIME") for _, v in vals)
                         or (len(lgs_here) == len([1 for _, v in vals
                                                   if v.upper() not in ("GT", "APT", "GAME TIME",
                                                                        "POSTING TIME", "LINE", "VS")]))):
            for col, v in vals:
                if v.upper() in ("GT", "GAME TIME"):
                    gt_col = col
                elif v.upper() in ("APT", "POSTING TIME"):
                    apt_col = col
            for col, lg in lgs_here:
                # league applies to the nearest category column at/left of it
                col_lg[col] = lg
            continue
        # pick row: emit any non-empty cell under a known category column
        gt = next((v for col, v in vals if col == gt_col), "")
        apt = next((v for col, v in vals if col == apt_col), "")
        for col, v in vals:
            if col in (gt_col, apt_col):
                continue
            if v.upper() in ("VS", "LINE", "-", "--"):
                continue
            # category: exact col or nearest category col to the left within 2
            cat = col_cat.get(col)
            src = col
            if cat is None:
                for d in (1, 2):
                    if col - d in col_cat:
                        cat = col_cat[col - d]; src = col - d
                        break
            if cat is None:
                continue
            # pure numeric cell = the LINE column, attach to nothing
            if re.fullmatch(r"[-+]?\d+(\.\d+)?", v):
                continue
            lg = col_lg.get(col) or col_lg.get(src)
            if lg is None:
                for d in (1, 2):
                    if col - d in col_lg:
                        lg = col_lg[col - d]
                        break
            out.append({"date": date, "league": lg or "", "category": cat,
                        "text": v[:120], "gt": gt[:24], "apt": apt[:24], "sheet_id": sid})
    wb.close()
    return len(out) - n0

def main():
    idx = json.load(open(SP + "/menu_archive_index_full.json"))
    date_of = {e["id"]: e["date"] for e in idx}
    files = sorted(os.listdir(RAW))
    out, bad = [], []
    for k, fn in enumerate(files):
        if not fn.endswith(".xlsx"):
            continue
        sid = fn[:-5]
        try:
            parse_book(f"{RAW}/{fn}", date_of.get(sid, "?"), sid, out)
        except Exception as ex:  # noqa: BLE001
            bad.append((sid, str(ex)[:80]))
        if (k + 1) % 100 == 0:
            print(f"  {k+1}/{len(files)} parsed, {len(out)} picks", flush=True)
    os.makedirs("data", exist_ok=True)
    with open("data/menu_picks.csv", "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["date", "league", "category", "text",
                                           "gt", "apt", "sheet_id"])
        w.writeheader()
        for r in sorted(out, key=lambda r: (r["date"], r["category"])):
            w.writerow(r)
    print(f"\n{len(out)} picks from {len(files)} files ({len(bad)} failed)")
    from collections import Counter
    print("by category:", dict(Counter(r['category'] for r in out).most_common()))
    print("by league:", dict(Counter(r['league'] for r in out).most_common()))
    print("by year:", dict(sorted(Counter(r['date'][:4] for r in out).items())))
    if bad:
        print("failed:", bad[:5])

if __name__ == "__main__":
    main()
